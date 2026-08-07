#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
把用户在 Excel 里回填的亚马逊联盟链接 + 本地下载的封面图，
更新进 src/data/deals.json 与 public/covers/。

映射逻辑：
  - Excel ②表 中 network=Amazon 的行，按出现顺序对应图片 YMX1..YMX15；
  - 每行按「游戏名称」匹配 deals.json 里 network=amazon 的同名 deal（取其 id）；
  - deals.json: 该 deal 的 storeUrl = amzn.to 联盟链接，image = /covers/{id}.jpg；
  - 图片: 复制 YMX{n}.jpg -> public/covers/{id}.jpg。
"""
import json
import shutil
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "ConsoleDeals_联盟链接表.xlsx"
DEALS = ROOT / "src/data/deals.json"
IMG_SRC = Path(r"C:\Users\LJH\Desktop\新建文件夹 (2)")
COVER_DST = ROOT / "public/covers"


def norm(s: str) -> str:
    return "".join(ch for ch in str(s).lower() if ch.isalnum())


def find_img(i: int) -> Path | None:
    # i=1,2 的文件带下划线后缀（YMX1_.jpg / YMX2_.jpg），其余为 YMX{n}.jpg
    for cand in (IMG_SRC / f"YMX{i}_.jpg", IMG_SRC / f"YMX{i}.jpg"):
        if cand.exists():
            return cand
    return None


def main():
    import openpyxl

    wb = openpyxl.load_workbook(XLSX)
    ws = wb["②游戏链接填报表"]

    # 取 Amazon 行（按 Excel 出现顺序），链接优先取第8列，回退第7列
    amazon_rows = []
    for r in range(4, ws.max_row + 1):
        net = ws.cell(row=r, column=2).value
        title = ws.cell(row=r, column=3).value
        link = ws.cell(row=r, column=8).value or ws.cell(row=r, column=7).value
        if net and "amazon" in str(net).lower() and link and "amzn.to" in str(link):
            amazon_rows.append((str(title).strip(), str(link).strip()))
    print(f"[1] Excel 中带 amzn.to 链接的 Amazon 行数: {len(amazon_rows)}")

    data = json.loads(DEALS.read_text(encoding="utf-8"))
    by_title = {
        norm(d["title"]): d for d in data["deals"] if d.get("network") == "amazon"
    }
    print(f"[2] deals.json 中 Amazon deal 数: {len(by_title)}")

    COVER_DST.mkdir(parents=True, exist_ok=True)

    report, ok, miss = [], 0, 0
    for i, (title, link) in enumerate(amazon_rows, 1):
        d = by_title.get(norm(title))
        if not d:
            report.append(f"  ✗ 未匹配标题: YMX{i} -> {title!r}")
            miss += 1
            continue
        did = d["id"]
        d["storeUrl"] = link  # 链接始终更新
        src = find_img(i)
        if not src:
            # 图片缺失：清掉 image 字段，让 DealCard 回退到 /covers/{id}.svg 占位
            d.pop("image", None)
            report.append(f"  ✗ 缺图片: YMX{i} ({title}) -> 目标 {did}.jpg（链接已更新，封面暂用占位图）")
            miss += 1
            continue
        d["image"] = f"/covers/{did}.jpg"
        shutil.copy(src, COVER_DST / f"{did}.jpg")
        report.append(f"  ✓ YMX{i} -> {did}.jpg | {title} | {link}")
        ok += 1

    DEALS.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"[3] 成功 {ok} / 缺失 {miss}")
    print("\n".join(report))


if __name__ == "__main__":
    main()
