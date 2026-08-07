#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
从 Excel『②游戏链接填报表』重建 src/data/deals.json。
- 图片按 Excel 行号映射：YMX1-15(亚马逊) / Green16-60(GMG) / Humble61-62 / Fanatical63-64
- 链接取第 8 列（【请回填】联盟推广链接），已是具体商品页或搜索页。
- 重建后所有 64 个游戏以 Excel 为准（GMG/Humble/Fanatical 部分游戏已被用户替换）。
"""
import json, re, shutil, os
from pathlib import Path
import openpyxl

ROOT = Path(r"C:\Users\LJH\Desktop\GameDealsSite")
XLSX = ROOT / "ConsoleDeals_联盟链接表.xlsx"
DEALS = ROOT / "src/data/deals.json"
IMG_SRC = Path(r"C:\Users\LJH\Desktop\新建文件夹 (2)")
COVER_DST = ROOT / "public/covers"
TODAY = "2026-08-07"

NET_MAP = {
    "amazon associates": "amazon",
    "green man gaming": "gmg",
    "humble bundle": "humble",
    "fanatical": "fanatical",
}
PREFIX = {"amazon": "YMX", "gmg": "Green", "humble": "Humble", "fanatical": "Fanatical"}

def slug(s: str) -> str:
    s = s.lower()
    # 去掉注册商标/商标符号、冒号、各种引号(含撇号)，避免生成带特殊字符的文件名/URL
    s = s.replace("®", "").replace("™", "").replace(":", " ")
    s = s.replace("’", "").replace("'", "").replace("”", "").replace("“", "").replace('"', "")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", "-", s.strip())
    return s

def find_img(prefix: str, k: int) -> Path | None:
    """按前缀+行号找源图（容忍尾随下划线/未知扩展名）。"""
    pat = f"{prefix}{k}*"
    hits = sorted(IMG_SRC.glob(pat))
    return hits[0] if hits else None

def fnum(v) -> float | None:
    """把 '$59.99' / '59.99' / '' 解析成 float，失败返回 None。"""
    if v is None:
        return None
    try:
        return float(str(v).replace("$", "").replace(",", "").strip())
    except Exception:
        return None

# 现有 deals（用于继承 note / lastChecked / productUrl）
existing = {d["id"]: d for d in json.loads(DEALS.read_text(encoding="utf-8"))["deals"]}

wb = openpyxl.load_workbook(XLSX)
ws = wb["②游戏链接填报表"]
hdr_row = None
for r in range(1, 6):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 12)]
    if any(v and "游戏名称" in str(v) for v in (vals or [])):
        hdr_row = r
        break

new_deals = []
report = []
missing_imgs = []

for idx, r in enumerate(range(hdr_row + 1, ws.max_row + 1), start=1):
    net_raw = ws.cell(row=r, column=2).value
    title = ws.cell(row=r, column=3).value
    platform = ws.cell(row=r, column=4).value or ""
    tags_raw = ws.cell(row=r, column=5).value or ""
    orig_raw = ws.cell(row=r, column=8).value
    cur_raw = ws.cell(row=r, column=9).value
    link = ws.cell(row=r, column=10).value or ""
    if not title or not link:
        continue
    net = NET_MAP.get(str(net_raw).strip().lower())
    if not net:
        report.append(f"  ✗ 无法识别网络: row{r} {net_raw!r}")
        continue
    title = str(title).strip()
    did = slug(title)
    tags = [t.strip() for t in str(tags_raw).split(",") if t.strip()]
    is_search = "search" in str(link).lower()
    url_status = "search" if is_search else "verified"
    orig_v = fnum(orig_raw)
    cur_v = fnum(cur_raw)

    # 图片
    prefix = PREFIX[net]
    src = find_img(prefix, idx)
    ext = src.suffix if src else None
    if src:
        dst = COVER_DST / f"{did}{ext}"
        shutil.copy(src, dst)
        image = f"/covers/{did}{ext}"
    else:
        image = None
        missing_imgs.append(f"YMX/Green/Humble/Fanatical idx={idx} ({title})")

    # 继承旧 note / lastChecked / productUrl
    old = existing.get(did, {})
    note = old.get("note") or (
        f"Compare editions and current price on {net_raw}. "
        f"Stock and regional pricing vary — verify on the store before buying."
    )
    last_checked = old.get("lastChecked") or TODAY
    product_url = old.get("productUrl")

    d = {
        "id": did,
        "title": title,
        "network": net,
        "platform": str(platform).strip(),
        "tags": tags,
        "storeUrl": str(link).strip(),
        "note": note,
        "lastChecked": last_checked,
        "urlStatus": url_status,
    }
    if image:
        d["image"] = image
    if product_url:
        d["productUrl"] = product_url
    if orig_v is not None:
        d["originalPrice"] = orig_v
    if cur_v is not None:
        d["price"] = cur_v
    new_deals.append(d)
    pr = "价:无" if orig_v is None else f"MSRP ${orig_v:g}" + (f" 现 ${cur_v:g}" if cur_v is not None else "")
    report.append(f"  ✓ [{idx:2d}] {net:9s} {title[:34]:34s} img={'有' if image else '缺'} {pr}")

# 写回 deals.json
data = {"deals": new_deals}
DEALS.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")

print(f"重建完成：共 {len(new_deals)} 个游戏")
print(f"缺图：{len(missing_imgs)} 个 -> {missing_imgs or '无'}")
print("-" * 90)
for line in report:
    print(line)

from collections import Counter
print("-" * 90)
print("网络分布:", dict(Counter(d["network"] for d in new_deals)))
print("搜索链接(未精确到商品页):", sum(1 for d in new_deals if d["urlStatus"] == "search"))
