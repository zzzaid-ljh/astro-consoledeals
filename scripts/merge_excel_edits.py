#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
把用户在 Excel『②游戏链接填报表』里的最新编辑合并进 deals.json。
策略：保留脚本此前抓取的 59 个价格，只覆盖用户本次改动的字段：
  - col10(联盟推广链接) -> storeUrl / productUrl（已是具体商品页/已修正的链接）
  - col8(MSRP)          -> originalPrice（修正捆绑包偏高 MSRP）
  - col9(现价,选填)     -> price（用户填了的才覆盖；其它保留原抓取价）
  - Amazon：按用户要求不显示具体价 -> price 统一置 None，卡片改显「Check price on Amazon」
  - urlStatus：col10 含 search 则 search，否则 verified
同时：对 3 个 GMG 中用户未填现价的（Halo / Ratchet）按新商品链接抓取现价补全。
"""
import json, re, html, ssl, time, random
from pathlib import Path
import openpyxl
import urllib.request

ROOT = Path(r"C:\Users\LJH\Desktop\GameDealsSite")
XLSX = ROOT / "ConsoleDeals_联盟链接表.xlsx"
DEALS = ROOT / "src/data/deals.json"
TODAY = "2026-08-07"
NET_MAP = {
    "amazon associates": "amazon",
    "green man gaming": "gmg",
    "humble bundle": "humble",
    "fanatical": "fanatical",
}

def norm_title(t):
    return html.unescape(str(t)).lower().strip()

def fnum(v):
    if v is None:
        return None
    try:
        return float(str(v).replace("$", "").replace(",", "").strip())
    except Exception:
        return None

# ---- 加载 deals.json ----
db = json.loads(DEALS.read_text(encoding="utf-8"))
deals = db["deals"]
by_title = {norm_title(d["title"]): d for d in deals}

# ---- 加载 Excel ----
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["②游戏链接填报表"]
hdr_row = None
for r in range(1, 6):
    vals = [ws.cell(row=r, column=c).value for c in range(1, 12)]
    if any(v and "游戏名称" in str(v) for v in (vals or [])):
        hdr_row = r
        break

excel_rows = []
for r in range(hdr_row + 1, ws.max_row + 1):
    net_raw = ws.cell(row=r, column=2).value
    title = ws.cell(row=r, column=3).value
    if not title:
        continue
    net = NET_MAP.get(str(net_raw).strip().lower())
    link = ws.cell(row=r, column=10).value or ""
    orig = fnum(ws.cell(row=r, column=8).value)
    cur = fnum(ws.cell(row=r, column=9).value)
    excel_rows.append((norm_title(title), net, link, orig, cur))

# ---- 合并 ----
log = []
for ntitle, net, link, orig, cur in excel_rows:
    d = by_title.get(ntitle)
    if not d:
        log.append(f"  ✗ 未匹配: {ntitle}")
        continue
    if link:
        d["storeUrl"] = link
        d["productUrl"] = link
    if orig is not None:
        d["originalPrice"] = orig
    if net == "amazon":
        d["price"] = None  # 用户要求亚马逊不显示具体价
    else:
        if cur is not None:
            d["price"] = cur  # 用户填了的现价
        elif "price" not in d:
            d["price"] = None  # 未抓到/未填 -> 显式置 None，保证字段存在
    d["urlStatus"] = "search" if "search" in link.lower() else "verified"
    log.append(f"  ✓ {d['id']:<42} net={net:<9} price={d.get('price')} orig={d.get('originalPrice')} status={d.get('urlStatus')}")

DEALS.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")
print("合并完成，逐条：")
for l in log:
    print(l)

# ---- 抓取用户未填现价的 GMG（Halo / Ratchet）----
def fetch_gmg_price(url):
    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"})
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE
        with urllib.request.urlopen(req, timeout=30, context=ctx) as resp:
            data = resp.read(400000).decode("utf-8", "ignore")
        m = re.search(r'"FormattedCurrentPrice"\s*:\s*"([0-9.,]+)"', data)
        if m:
            return float(m.group(1).replace(",", ""))
        m2 = re.search(r'"price"\s*:\s*"([0-9.]+)"', data)
        if m2:
            return float(m2.group(1))
    except Exception as e:
        print(f"  ! fetch失败 {url}: {e}")
    return None

targets = ["halo-infinite", "ratchet-38-clank-rift-apart"]
print("\n抓取未填现价的 GMG：")
db2 = json.loads(DEALS.read_text(encoding="utf-8"))
for d in db2["deals"]:
    if d["id"] in targets and d.get("price") is None:
        p = fetch_gmg_price(d["storeUrl"])
        if p:
            d["price"] = p
            d["lastChecked"] = TODAY
            print(f"  ✓ {d['id']} -> ${p}")
        else:
            print(f"  ✗ {d['id']} 抓取失败，保留 MSRP 显示")
        time.sleep(random.uniform(1.0, 2.0))
DEALS.write_text(json.dumps(db2, ensure_ascii=False, indent=2), encoding="utf-8")
print("\n完成。")
