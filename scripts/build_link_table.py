#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""从 deals.json 生成「联盟链接填报表」xlsx，供用户申请联盟身份后回填推广链接。"""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(r"C:\Users\LJH\Desktop\GameDealsSite")
DEALS = json.loads((ROOT / "src/data/deals.json").read_text(encoding="utf-8"))["deals"]

NET_DISPLAY = {
    "amazon": "Amazon Associates",
    "gmg": "Green Man Gaming",
    "humble": "Humble Bundle",
    "fanatical": "Fanatical",
}
NET_ORDER = {"amazon": 0, "gmg": 1, "humble": 2, "fanatical": 3}

def msrp_for(platform: str) -> float:
    """按平台预填参考售价（USD）。仅作起点，用户可在 Excel 里核对/修改。"""
    p = str(platform).lower()
    if "switch 2" in p:
        return 69.99
    if "switch" in p:
        return 59.99
    if "playstation 5" in p or "ps5" in p:
        return 69.99
    if "xbox" in p:
        return 69.99
    if "pc" in p or "steam" in p:
        return 59.99
    return 59.99

HEADER_FILL = PatternFill("solid", fgColor="15A06B")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
FILL_BACK = PatternFill("solid", fgColor="FFF4CC")   # 回填列高亮
TITLE_FONT = Font(bold=True, size=14, color="1C2530")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="D0D5DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()

# ---------- Sheet 1: 联盟申请清单 ----------
ws1 = wb.active
ws1.title = "①联盟申请清单"
ws1["A1"] = "联盟身份申请清单（先申请这几个，拿到联盟 ID 后再填游戏链接）"
ws1["A1"].font = TITLE_FONT
ws1.merge_cells("A1:G1")

app_rows = [
    ["Amazon Associates", "亚马逊官方联盟", "https://affiliate-program.amazon.com/",
     "数字游戏 1-2%", "24h cookie；180天内需出3单否则封号；强制FTC声明；禁止手标价格/用短链",
     "（填：已申请/已通过/待填ID）",
     "先不急，等站内容厚了再申，避免薄内容被拒"],
    ["Green Man Gaming", "Tapfiliate 平台", "https://www.greenmangaming.com/affiliates/",
     "5-10%", "30天 cookie；数字 key 为主，转化稳",
     "（填：已申请/已通过/待填ID）",
     "本表多数 PS5/Xbox/跨平台游戏走这家"],
    ["Humble Bundle", "Impact 平台", "https://www.humblebundle.com/affiliates",
     "约 5-10%", "30天 cookie；商店与捆绑包都带联盟",
     "（填：已申请/已通过/待填ID）",
     "本表 EA FC 25、博德之门3 走这家"],
    ["Fanatical", "官方联盟", "https://www.fanatical.com/en/affiliates",
     "5-10%", "30天 cookie；清仓折扣力度大",
     "（填：已申请/已通过/待填ID）",
     "本表 霍格沃茨legacy、消逝的光芒2 走这家"],
]
hdr1 = ["网站", "联盟性质/平台", "申请网址", "大概佣金", "关键政策（Cookie等）", "申请状态（回填）", "备注"]
ws1.append([])  # row2 spacer
ws1.append(hdr1)
hrow = ws1.max_row
for c in range(1, len(hdr1) + 1):
    cell = ws1.cell(row=hrow, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = BORDER
for r in app_rows:
    ws1.append(r)
    rr = ws1.max_row
    for c in range(1, len(hdr1) + 1):
        cell = ws1.cell(row=rr, column=c)
        cell.alignment = WRAP
        cell.border = BORDER
widths1 = [20, 18, 42, 14, 40, 22, 30]
for i, w in enumerate(widths1, start=1):
    ws1.column_dimensions[chr(64 + i)].width = w
ws1.freeze_panes = "A4"

# ---------- Sheet 2: 游戏链接填报表 ----------
ws2 = wb.create_sheet("②游戏链接填报表")
ws2["A1"] = "游戏 → 联盟推广链接填报表（黄底列请你回填，然后整个文件发回给我）"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:H1")

hdr2 = ["序号", "网站/联盟", "游戏名称", "平台", "标签",
        "当前商品链接（具体商品页）", "链接状态",
        "原价/MSRP(USD,参考)", "现价(USD,选填)", "【请回填】联盟推广链接", "回填状态"]
ws2.append([])
ws2.append(hdr2)
hrow2 = ws2.max_row
for c in range(1, len(hdr2) + 1):
    cell = ws2.cell(row=hrow2, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = BORDER

# 按 network 分组排序，方便你针对同一家网站批量申请
# 链接状态配色：predicted 用浅橙提醒用户确认，verified 正常
FILL_PREDICT = PatternFill("solid", fgColor="FCE4D6")
deals_sorted = sorted(DEALS, key=lambda d: (NET_ORDER.get(d["network"], 9), d["title"]))
for i, d in enumerate(deals_sorted, start=1):
    net = d["network"]
    prod_url = d.get("productUrl", d["storeUrl"])
    url_status = d.get("urlStatus", "unknown")
    status_label = {"verified": "已核实", "predicted": "推测·请确认", "unknown": "未知"}.get(url_status, url_status)
    row = [
        i,
        NET_DISPLAY.get(net, net),
        d["title"],
        d["platform"],
        ", ".join(d.get("tags", [])),
        prod_url,
        status_label,
        d.get("originalPrice", msrp_for(d["platform"])),   # 原价 / MSRP（参考）
        d.get("price", ""),                                # 现价（选填）
        d.get("storeUrl", ""),   # 回填列（保留已有联盟链接，避免重生成丢数据）
        "",   # 回填状态
    ]
    ws2.append(row)
    rr = ws2.max_row
    for c in range(1, len(hdr2) + 1):
        cell = ws2.cell(row=rr, column=c)
        cell.border = BORDER
        if c == 10:  # 回填列（黄底高亮）
            cell.fill = FILL_BACK
            cell.alignment = WRAP
        elif c == 6:  # 当前商品链接
            cell.alignment = WRAP
        elif c == 7:  # 链接状态
            cell.alignment = CENTER
            if url_status == "predicted":
                cell.fill = FILL_PREDICT
        elif c in (3, 5, 8, 9):  # 游戏名称、标签、价格列
            cell.alignment = WRAP
        else:  # 序号/网站/平台/回填状态
            cell.alignment = CENTER

widths2 = [6, 20, 38, 20, 22, 50, 14, 16, 16, 50, 12]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[chr(64 + i)].width = w
ws2.freeze_panes = "A4"
ws2.auto_filter.ref = f"A{hrow2}:K{ws2.max_row}"

# ---------- Sheet 3: 填写说明 ----------
ws3 = wb.create_sheet("③填写说明")
ws3["A1"] = "怎么用这张表"
ws3["A1"].font = TITLE_FONT
guide = [
    "",
    "1) 先看【①联盟申请清单】 sheet，按里面网址去申请 4 个联盟账号。",
    "2) 申请通过后，点开本表『当前商品链接（具体商品页）』列里那行对应的链接，在网站里找到这个商品，用你的联盟后台生成『带你追踪ID的推广链接』。",
    "3) 回到【②游戏链接填报表】，在黄底列『【请回填】联盟推广链接』粘贴该链接；『回填状态』写『已填』。",
    "4) 不用每个都填：先填你已通过的那几家网站对应的行即可，其余留空。",
    "5) 填完后把整个 .xlsx 文件发回给我，我会把链接写进网站（更新 deals.json 的 storeUrl，并按需打开 config.ts 里对应网络的 enabled）。",
    "",
    "关于『链接状态』列：",
    "- 『已核实』：商品页 URL 已逐个联网核实，点开就是正确商品，直接用它生成联盟链接即可。",
    "- 『推测·请确认』（橙色单元格）：GMG（被反爬保护，无法自动核实）和 Humble（官方商品页未搜到）这两家给的是『按命名规律推测的链接』，点开若是 404 或跳首页，请在该网站搜索框里搜游戏名，用你搜到的真实商品页生成联盟链接。",
    "",
    "价格列说明：",
    "- 『原价/MSRP(USD,参考)』：我按平台预填了参考售价（Switch $59.99 / Switch 2·PS5·Xbox $69.99），你可按实际核对修改。",
    "- 『现价(USD,选填)』：留空即可；等你拿到真实促销价再填，网站会自动算出并显示『Save X%』折扣徽章。",
    "- 展示的价格需与实际一致，长期建议用 Amazon Product Advertising API（PA-API）自动同步，避免过期价导致合规风险。",
    "",
    "注意事项：",
    "- 亚马逊：别用短链（bit.ly等）、文章页必须有 FTC 声明（网站模板已内置）。",
    "- 一个游戏当前只挂了一个网站（见『网站/联盟』列），所以每行只需填那一家对应的推广链接。",
    "- 若你某游戏想同时挂多家（比如 GMG + Humble），告诉我，我可以把那行拆成多行。",
    "- 推广链接必须是『带你自己联盟ID』的完整链接，否则佣金不算你的。",
]
for line in guide:
    ws3.append([line])
ws3.column_dimensions["A"].width = 100
for r in range(2, ws3.max_row + 1):
    ws3.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")

out = ROOT / "ConsoleDeals_联盟链接表.xlsx"
wb.save(out)
print("SAVED:", out)
print("游戏行数:", len(deals_sorted))
print("网站分布:", {k: sum(1 for d in DEALS if d["network"] == k) for k in NET_ORDER})
